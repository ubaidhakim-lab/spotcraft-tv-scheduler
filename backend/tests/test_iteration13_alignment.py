"""Iteration 13: Wk banner + Day-name row alignment with Date row in downloaded xlsx.

Bug: build_output_workbook had off-by-one shifting Wk banner and Day-name rows one column
to the right of date columns. Fix removed the extra '+ 1'. This suite verifies the invariant
across multiple campaign start dates.
"""
import io
import os
import datetime as dt

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get(
    'REACT_APP_BACKEND_URL', 'https://spot-scheduling-hub.preview.emergentagent.com'
).rstrip('/')
API = f"{BASE_URL}/api"
SAMPLE = "/tmp/samples/input.xlsx"

DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


@pytest.fixture(scope="module")
def plan():
    with open(SAMPLE, "rb") as f:
        files = {"file": ("input.xlsx", f,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{API}/plans/upload", files=files, timeout=60)
    assert r.status_code == 200, r.text
    return r.json()


def _generate(plan, start, end):
    payload = {
        "edits": plan.get("default_edits") or [
            {"duration": 10, "percentage": 60},
            {"duration": 20, "percentage": 30},
            {"duration": 30, "percentage": 10},
        ],
        "row_overrides": [],
        "prefs": {
            "campaign_start": start,
            "campaign_end": end,
            "spot_frequency_minutes": 30,
            "movies_frequency_minutes": 60,
            "movies_genres": ["MOV", "Movies", "Movie"],
            "gec_genres": ["GEC"],
            "weekly_grp_dispersion": [],
            "blackout_days": [],
            "blackout_dates": [],
            "daypart_weights": [],
            "weekend_boost": 1.0,
            "reach_vs_frequency": 0.5,
        },
    }
    r = requests.post(f"{API}/plans/{plan['plan_id']}/generate", json=payload, timeout=180)
    assert r.status_code == 200, r.text
    return r.json()


def _download_ws(result_id):
    r = requests.get(f"{API}/results/{result_id}/download", timeout=60)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content), data_only=False)
    assert "Schedule Sheet" in wb.sheetnames, wb.sheetnames
    return wb["Schedule Sheet"]


def _find_header_row(ws):
    """Header row = the row containing 'Direct/Matrix' in col 1 (per problem statement)."""
    for row in range(1, min(50, ws.max_row + 1)):
        v = ws.cell(row=row, column=1).value
        if v and "Direct/Matrix" in str(v):
            return row
    # Fallback: any row starting a column containing 'Channel'/'Program' etc
    raise AssertionError("Header row with 'Direct/Matrix' not found in Schedule Sheet")


def _first_date_col(ws, header_row):
    """Find first column where header row cell is a date."""
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if isinstance(v, (dt.datetime, dt.date)):
            return col
    raise AssertionError("No date column found in header row")


def _check_alignment(ws, start_date, end_date):
    header_row = _find_header_row(ws)
    day_row = header_row - 1
    wk_row = header_row - 2
    assert day_row >= 1 and wk_row >= 1

    first_col = _first_date_col(ws, header_row)
    n_days = (end_date - start_date).days + 1
    weeks = (n_days + 6) // 7

    # For each day column verify: date row = start+i, day row = day-name
    for i in range(n_days):
        col = first_col + i
        dv = ws.cell(row=header_row, column=col).value
        if isinstance(dv, dt.datetime):
            dv = dv.date()
        expected_date = start_date + dt.timedelta(days=i)
        assert dv == expected_date, (
            f"Date col mismatch at i={i}: got {dv}, expected {expected_date}"
        )
        day_val = ws.cell(row=day_row, column=col).value
        expected_day = DAY_ORDER[expected_date.weekday()]
        assert day_val == expected_day, (
            f"Day name mismatch at col {col} (date {expected_date}): "
            f"got {day_val!r}, expected {expected_day!r}"
        )

    # Wk banner: for each week w, the column at first_col + w*7 should have 'Wk {w+1}'
    for w in range(weeks):
        col = first_col + w * 7
        wk_val = ws.cell(row=wk_row, column=col).value
        assert wk_val == f"Wk {w+1}", (
            f"Wk banner mismatch at week {w+1} col {col}: got {wk_val!r}"
        )

    # Also assert NO Wk label appears at any non-week-start column in that row within the date range
    for i in range(n_days):
        col = first_col + i
        wk_val = ws.cell(row=wk_row, column=col).value
        if i % 7 == 0:
            assert wk_val == f"Wk {i//7 + 1}"
        else:
            assert wk_val in (None, ""), (
                f"Unexpected Wk banner at offset i={i} col {col}: {wk_val!r}"
            )


class TestAlignmentAugSaturdayStart:
    """Primary: 2026-08-01 (Sat) — the reported bug case."""

    def test_alignment(self, plan):
        res = _generate(plan, "2026-08-01", "2026-08-31")
        ws = _download_ws(res["result_id"])
        _check_alignment(ws, dt.date(2026, 8, 1), dt.date(2026, 8, 31))


class TestAlignmentJuneMondayStart:
    """Arbitrary Monday start."""

    def test_alignment(self, plan):
        res = _generate(plan, "2026-06-29", "2026-08-16")
        ws = _download_ws(res["result_id"])
        _check_alignment(ws, dt.date(2026, 6, 29), dt.date(2026, 8, 16))


class TestAlignmentWednesdayStart:
    """Arbitrary Wednesday start for extra coverage."""

    def test_alignment(self, plan):
        res = _generate(plan, "2026-09-02", "2026-09-29")
        ws = _download_ws(res["result_id"])
        _check_alignment(ws, dt.date(2026, 9, 2), dt.date(2026, 9, 29))
