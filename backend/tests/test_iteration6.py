"""Iteration 6 tests: strict campaign_end enforcement (partial last week)."""
import os
import io
import openpyxl
import pytest
import requests
from datetime import date, timedelta

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://spot-scheduling-hub.preview.emergentagent.com",
).rstrip("/")
SAMPLE = "/tmp/samples/input.xlsx"


def _upload():
    with open(SAMPLE, "rb") as fh:
        r = requests.post(
            f"{BASE_URL}/api/plans/upload",
            files={"file": ("input.xlsx", fh,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 200, r.text
    return r.json()["plan_id"]


def _generate(pid, campaign_start, campaign_end, weeks):
    payload = {
        "edits": [
            {"duration": 30, "percentage": 50},
            {"duration": 20, "percentage": 20},
            {"duration": 45, "percentage": 20},
            {"duration": 10, "percentage": 10},
        ],
        "prefs": {
            "campaign_start": campaign_start,
            "campaign_end": campaign_end,
            "campaign_weeks": weeks,
            "spot_frequency_minutes": 30,
            "gec_genres": ["GEC"],
            "weekly_grp_dispersion": [100 // weeks] * weeks,
            "blackout_days": [],
            "blackout_dates": [],
            "daypart_weights": [],
        },
        "row_overrides": [],
    }
    r = requests.post(f"{BASE_URL}/api/plans/{pid}/generate", json=payload)
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="module")
def partial_result():
    pid = _upload()
    # 2026-06-29 (Mon) → 2026-08-12 (Wed) = 45 days = 6 full weeks + 3 days
    res = _generate(pid, "2026-06-29", "2026-08-12", 7)
    return res


@pytest.fixture(scope="module")
def full_result():
    pid = _upload()
    # 2026-06-29 (Mon) → 2026-08-16 (Sun) = 49 days = 7 full weeks
    res = _generate(pid, "2026-06-29", "2026-08-16", 7)
    return res


# ---------- STRICT END: schedule_rows ----------
def test_partial_no_dates_past_end(partial_result):
    rows = partial_result.get("schedule_rows", [])
    assert rows, "schedule_rows empty"
    end = date(2026, 8, 12)
    start = date(2026, 6, 29)
    dates = set()
    for r in rows:
        d = r.get("date") or r.get("Date")
        assert d, f"row missing date: {r}"
        dd = date.fromisoformat(d[:10])
        assert start <= dd <= end, f"date {dd} outside campaign window"
        dates.add(dd)
    # (b) at least one row on last date
    assert end in dates, "no schedule row on campaign_end (2026-08-12)"
    # (c) unique dates count == 45
    assert len(dates) == 45, f"expected 45 unique dates, got {len(dates)}"


# ---------- STRICT END: downloaded xlsx ----------
def test_partial_download_date_columns(partial_result):
    rid = partial_result["result_id"]
    r = requests.get(f"{BASE_URL}/api/results/{rid}/download")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["Schedule Sheet"]

    # find header row containing 'channel'
    hdr_row_idx = None
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and any(str(c or "").strip().lower() == "channel" for c in row):
            hdr_row_idx = i
            header = list(row)
            break
    assert hdr_row_idx, "header row not found"

    # daily date columns: cells that parse as a date
    date_cols = []
    for c in header:
        if isinstance(c, date):
            date_cols.append(c if not hasattr(c, "date") else c)
        else:
            # try parse "YYYY-MM-DD"
            if isinstance(c, str) and len(c) >= 10:
                try:
                    date_cols.append(date.fromisoformat(c[:10]))
                except ValueError:
                    pass
    # normalize to date objects
    norm = []
    for d in date_cols:
        norm.append(d.date() if hasattr(d, "date") else d)
    assert len(norm) == 45, f"expected 45 date columns, got {len(norm)}: {norm[:3]}...{norm[-3:]}"
    assert norm[0] == date(2026, 6, 29), f"first date column {norm[0]} != 2026-06-29"
    assert norm[-1] == date(2026, 8, 12), f"last date column {norm[-1]} != 2026-08-12"
    assert date(2026, 8, 13) not in norm
    assert date(2026, 8, 16) not in norm


# ---------- PARTIAL WEEK 7 exists ----------
def test_partial_week7_present(partial_result):
    rid = partial_result["result_id"]
    r = requests.get(f"{BASE_URL}/api/results/{rid}/download")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["Schedule Sheet"]
    strings = []
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if isinstance(c, str):
                strings.append(c.lower())
    txt = " | ".join(strings)
    assert "wk 7" in txt, "Week 7 header not present in partial campaign"
    assert "wk 1" in txt


# ---------- REGRESSION: full-week ----------
def test_full_week_49_days(full_result):
    rows = full_result.get("schedule_rows", [])
    assert rows
    end = date(2026, 8, 16)
    start = date(2026, 6, 29)
    dates = set()
    for r in rows:
        d = r.get("date") or r.get("Date")
        dd = date.fromisoformat(d[:10])
        assert start <= dd <= end
        dates.add(dd)
    assert end in dates, "missing last date 2026-08-16"
    assert len(dates) == 49, f"expected 49 unique dates, got {len(dates)}"


def test_full_week_download_49_cols(full_result):
    rid = full_result["result_id"]
    r = requests.get(f"{BASE_URL}/api/results/{rid}/download")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["Schedule Sheet"]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and any(str(c or "").strip().lower() == "channel" for c in row):
            header = list(row)
            break
    dates_seen = []
    for c in header:
        if hasattr(c, "date"):
            dates_seen.append(c.date() if hasattr(c, "date") and callable(getattr(c, "date", None)) else c)
        elif isinstance(c, str) and len(c) >= 10:
            try:
                dates_seen.append(date.fromisoformat(c[:10]))
            except ValueError:
                pass
    norm = [d.date() if hasattr(d, "date") and callable(getattr(d, "date", None)) else d for d in dates_seen]
    # dedupe consecutive if any
    assert len(norm) == 49, f"expected 49 columns, got {len(norm)}"
    assert norm[0] == date(2026, 6, 29)
    assert norm[-1] == date(2026, 8, 16)


# ---------- REGRESSION: blackout_dates never pushes past end ----------
def test_blackout_does_not_push_past_end():
    pid = _upload()
    payload = {
        "edits": [
            {"duration": 30, "percentage": 50},
            {"duration": 20, "percentage": 20},
            {"duration": 45, "percentage": 20},
            {"duration": 10, "percentage": 10},
        ],
        "prefs": {
            "campaign_start": "2026-06-29",
            "campaign_end": "2026-08-12",
            "campaign_weeks": 7,
            "spot_frequency_minutes": 30,
            "gec_genres": ["GEC"],
            "weekly_grp_dispersion": [15, 15, 15, 15, 15, 15, 10],
            "blackout_days": [],
            # blackout the last 2 days
            "blackout_dates": ["2026-08-11", "2026-08-12"],
            "daypart_weights": [],
        },
        "row_overrides": [],
    }
    r = requests.post(f"{BASE_URL}/api/plans/{pid}/generate", json=payload)
    assert r.status_code == 200, r.text
    rows = r.json().get("schedule_rows", [])
    end = date(2026, 8, 12)
    for row in rows:
        d = row.get("date") or row.get("Date")
        dd = date.fromisoformat(d[:10])
        assert dd <= end, f"row scheduled past campaign_end even with blackouts: {dd}"
        assert dd not in (date(2026, 8, 11), date(2026, 8, 12)), \
            f"row scheduled on blacked-out date {dd}"
