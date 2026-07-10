"""Iteration 16 backend tests.

Covers latest fix: 'Do not drop spots - accommodate all planned spots'.
Also exhaustively covers upload/generate/download/sessions endpoints,
blackouts (dates + day-of-week), campaign_end respect, GEC even distribution,
movies stacking, FCT preservation, weekly_grp_dispersion, weekend_boost,
reach_vs_frequency, gec_planning_weeks.

All test data is synthesised in-memory via openpyxl - no fixture files on disk.
"""
import io
import os
import math
from collections import Counter, defaultdict
from datetime import datetime, timedelta

import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"


# ---------- helpers to build synthetic plan workbooks ----------

STANDARD_HEADERS = [
    "Market", "Genre", "Channel", "Program", "Days", "Start Time", "End Time",
    "Net Rate 10Sec", "ACD", "Spots", "FCT", "Net Outlay", "Log TVR", "GRP",
]

def build_plan_xlsx(rows):
    """rows: list of dicts keyed by STANDARD_HEADERS (missing keys -> blank)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    # metadata rows (2 rows) - upload endpoint tolerates any leading metadata
    ws.append(["Client", "TEST"])
    ws.append(["Campaign", "ITER16"])
    ws.append([])  # blank row
    ws.append(STANDARD_HEADERS)
    for r in rows:
        ws.append([r.get(h, "") for h in STANDARD_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def upload_plan(rows):
    buf = build_plan_xlsx(rows)
    resp = requests.post(
        f"{API}/plans/upload",
        files={"file": ("plan.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


def generate(plan_id, edits, prefs, row_overrides=None):
    body = {
        "edits": edits,
        "row_overrides": row_overrides or [],
        "prefs": prefs,
    }
    resp = requests.post(f"{API}/plans/{plan_id}/generate", json=body, timeout=60)
    assert resp.status_code == 200, resp.text
    return resp.json()


# ---------- 1. Upload parses standard columns ----------

class TestUpload:
    def test_upload_parses_all_standard_columns(self):
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "Colors Kannada",
            "Program": "Show A", "Days": "Mon-Fri",
            "Start Time": "20:00", "End Time": "22:00",
            "Net Rate 10Sec": 5000, "ACD": 20, "Spots": 30, "FCT": 600,
            "Net Outlay": 300000, "Log TVR": 3.5, "GRP": 105,
        }]
        data = upload_plan(rows)
        assert "plan_id" in data
        assert data["row_count"] == 1
        pr = data["rows"][0]
        assert pr["market"] == "KAR"
        assert pr["genre"] == "GEC"
        assert pr["channel"] == "Colors Kannada"
        assert pr["program"] == "Show A"
        assert pr["days"] == "Mon-Fri"
        assert float(pr["net_rate_10s"]) == 5000
        assert float(pr["acd"]) == 20
        assert float(pr["spots"]) == 30
        assert float(pr["fct"]) == 600
        assert float(pr["log_tvr"]) == 3.5
        assert float(pr["grp"]) == 105


# ---------- 2. Core invariant: NO SPOT DROPPED ----------

class TestNoSpotDropped:
    """Sum of edit_rows.final_spots must exactly equal len(schedule_rows)."""

    def _invariant(self, gen):
        er = gen["edit_rows"]
        sr = gen["schedule_rows"]
        total_final = sum(e["final_spots"] for e in er)
        assert len(sr) == total_final, (
            f"MISMATCH: schedule_rows={len(sr)} vs sum(final_spots)={total_final}"
        )
        # Per edit_row
        by_key = Counter((s["_row_id"], s["edit_duration"]) for s in sr)
        for e in er:
            k = (e["_row_id"], e["edit_duration"])
            assert by_key[k] == e["final_spots"], (
                f"per-edit mismatch row={k} sched={by_key[k]} final={e['final_spots']}"
            )

    def test_simple_gec(self):
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "CH1", "Program": "P1",
            "Days": "Mon-Fri", "Start Time": "20:00", "End Time": "22:00",
            "Net Rate 10Sec": 1000, "ACD": 20, "Spots": 60, "FCT": 1200,
        }]
        d = upload_plan(rows)
        gen = generate(d["plan_id"],
                       edits=[{"duration": 30, "percentage": 60}, {"duration": 20, "percentage": 30}, {"duration": 10, "percentage": 10}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-01-18"})
        self._invariant(gen)

    def test_stress_stacking_over_capacity(self):
        # 500 spots in a 1-hour band over 14 days => must stack.
        rows = [{
            "Market": "KAR", "Genre": "Movies", "Channel": "MoviesCh", "Program": "MovieBlock",
            "Days": "Mon-Sun", "Start Time": "23:00", "End Time": "24:00",
            "Net Rate 10Sec": 500, "ACD": 10, "Spots": 500, "FCT": 5000,
        }]
        d = upload_plan(rows)
        gen = generate(d["plan_id"],
                       edits=[{"duration": 10, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-01-18",
                              "movies_frequency_minutes": 60})
        self._invariant(gen)
        sr = gen["schedule_rows"]
        assert len(sr) == 500
        # Distribution across 14 days should be roughly even (35-36 per day)
        per_day = Counter(s["date"] for s in sr)
        assert len(per_day) == 14, f"expected 14 days, got {len(per_day)}"
        mx, mn = max(per_day.values()), min(per_day.values())
        assert mx - mn <= 2, f"uneven distribution min={mn} max={mx}"

    def test_multi_row_multi_edit(self):
        rows = [
            {"Market": "KAR", "Genre": "GEC", "Channel": f"CH{i}", "Program": f"P{i}",
             "Days": "Mon-Sun", "Start Time": "18:00", "End Time": "22:00",
             "Net Rate 10Sec": 1000, "ACD": 15, "Spots": 40, "FCT": 600}
            for i in range(1, 6)
        ]
        d = upload_plan(rows)
        gen = generate(d["plan_id"],
                       edits=[{"duration": 30, "percentage": 50}, {"duration": 10, "percentage": 50}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-02-01"})
        self._invariant(gen)


# ---------- 3. GEC: FCT preserved, even distribution, day/time bounds ----------

class TestGECEvenDistribution:
    def test_gec_prime_time_100plus_spots(self):
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "GECch", "Program": "GECprog",
            "Days": "Mon-Fri", "Start Time": "20:00", "End Time": "22:00",
            "Net Rate 10Sec": 2000, "ACD": 22, "Spots": 100, "FCT": 2200,
        }]
        d = upload_plan(rows)
        edits = [{"duration": 30, "percentage": 60},
                 {"duration": 20, "percentage": 30},
                 {"duration": 10, "percentage": 10}]
        gen = generate(d["plan_id"], edits=edits,
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-02-01"})
        er = gen["edit_rows"]
        sr = gen["schedule_rows"]
        # (a) FCT preservation: sum(dur * spots) == original acd * original_spots
        original_fct = 22 * 100  # 2200
        total_fct = sum(e["edit_duration"] * e["final_spots"] for e in er)
        # We accept off-by-a-few due to rounding of edit shares.
        assert abs(total_fct - original_fct) <= max(20, original_fct * 0.02), (
            f"FCT drift: got {total_fct} vs {original_fct}"
        )
        # (b) Even day distribution: max/min ratio < 1.5
        per_day = Counter(s["date"] for s in sr)
        assert per_day, "no schedule rows produced"
        mx, mn = max(per_day.values()), min(per_day.values())
        if mn > 0:
            ratio = mx / mn
            assert ratio < 1.6, f"uneven days ratio={ratio} min={mn} max={mx}"
        # (c) No Sat/Sun
        for s in sr:
            assert s["day"] not in ("Sat", "Sun"), f"Sat/Sun leaked: {s}"
        # (d) No time outside 20:00-22:00
        for s in sr:
            hh, mm = map(int, s["spot_time"].split(":"))
            assert 20 * 60 <= hh * 60 + mm < 22 * 60, f"time out of band: {s['spot_time']}"


# ---------- 4. Campaign end strict ----------

class TestCampaignEndRespected:
    def test_no_row_past_end(self):
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "CH", "Program": "P",
            "Days": "Mon-Sun", "Start Time": "18:00", "End Time": "20:00",
            "Net Rate 10Sec": 1000, "ACD": 20, "Spots": 50, "FCT": 1000,
        }]
        d = upload_plan(rows)
        end = "2026-01-18"
        gen = generate(d["plan_id"],
                       edits=[{"duration": 20, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": end})
        end_d = datetime.strptime(end, "%Y-%m-%d").date()
        for s in gen["schedule_rows"]:
            dt = datetime.strptime(s["date"], "%Y-%m-%d").date()
            assert dt <= end_d, f"schedule after campaign_end: {s['date']}"
            assert dt >= datetime.strptime("2026-01-05", "%Y-%m-%d").date()


# ---------- 5. Blackout dates + day-of-week ----------

class TestBlackouts:
    def test_blackout_dates_absent(self):
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "CH", "Program": "P",
            "Days": "Mon-Sun", "Start Time": "18:00", "End Time": "22:00",
            "Net Rate 10Sec": 1000, "ACD": 20, "Spots": 60, "FCT": 1200,
        }]
        d = upload_plan(rows)
        blackouts = ["2026-01-07", "2026-01-14"]
        gen = generate(d["plan_id"],
                       edits=[{"duration": 20, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-01-18",
                              "blackout_dates": blackouts})
        dates = {s["date"] for s in gen["schedule_rows"]}
        for bd in blackouts:
            assert bd not in dates, f"blackout date {bd} present"

    def test_blackout_day_of_week_absent(self):
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "CH", "Program": "P",
            "Days": "Mon-Sun", "Start Time": "18:00", "End Time": "22:00",
            "Net Rate 10Sec": 1000, "ACD": 20, "Spots": 60, "FCT": 1200,
        }]
        d = upload_plan(rows)
        gen = generate(d["plan_id"],
                       edits=[{"duration": 20, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-01-18",
                              "blackout_days": ["Sun"]})
        for s in gen["schedule_rows"]:
            assert s["day"] != "Sun", f"Sun leaked: {s}"


# ---------- 6. Excel download parseable ----------

class TestExcelDownload:
    def test_download_valid_excel_with_totals(self):
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "CH", "Program": "P",
            "Days": "Mon-Fri", "Start Time": "20:00", "End Time": "22:00",
            "Net Rate 10Sec": 1000, "ACD": 20, "Spots": 40, "FCT": 800,
        }]
        d = upload_plan(rows)
        gen = generate(d["plan_id"],
                       edits=[{"duration": 20, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-01-18"})
        total_spots = sum(e["final_spots"] for e in gen["edit_rows"])
        resp = requests.get(f"{API}/results/{gen['result_id']}/download", timeout=30)
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "").lower()
        wb = load_workbook(io.BytesIO(resp.content), data_only=True)
        ws = wb["Schedule Sheet"]
        # Find GRAND TOTAL row
        grand = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and "GRAND TOTAL" in str(v):
                grand = r
                break
        assert grand is not None, "GRAND TOTAL row missing"
        # Locate 'Final Spots' in headers
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Market":
                header_row = r
                break
        assert header_row is not None
        final_spots_col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=c).value == "Final Spots":
                final_spots_col = c
                break
        assert final_spots_col is not None
        grand_spots = ws.cell(row=grand, column=final_spots_col).value
        assert grand_spots == total_spots, f"grand={grand_spots} vs {total_spots}"


# ---------- 7. Sessions CRUD ----------

class TestSessions:
    def test_session_crud(self):
        # need a plan_id
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "CH", "Program": "P",
            "Days": "Mon-Fri", "Start Time": "20:00", "End Time": "22:00",
            "Net Rate 10Sec": 1000, "ACD": 20, "Spots": 10, "FCT": 200,
        }]
        d = upload_plan(rows)
        body = {
            "name": "TEST_session_iter16",
            "plan_id": d["plan_id"],
            "edits": [{"duration": 20, "percentage": 100}],
            "row_overrides": [],
            "prefs": {"campaign_start": "2026-01-05", "campaign_end": "2026-01-18"},
        }
        r = requests.post(f"{API}/sessions", json=body, timeout=15)
        assert r.status_code == 200, r.text
        sid = r.json()["id"]
        # list
        lst = requests.get(f"{API}/sessions", timeout=15).json()
        assert any(s["id"] == sid for s in lst)
        # get
        g = requests.get(f"{API}/sessions/{sid}", timeout=15)
        assert g.status_code == 200
        assert g.json()["name"] == "TEST_session_iter16"
        # delete
        dl = requests.delete(f"{API}/sessions/{sid}", timeout=15)
        assert dl.status_code == 200
        g2 = requests.get(f"{API}/sessions/{sid}", timeout=15)
        assert g2.status_code == 404


# ---------- 8. Extra prefs: weekly_grp_dispersion, weekend_boost, reach_vs_frequency, gec_planning_weeks ----------

class TestAdvancedPrefs:
    def _make_plan(self):
        rows = [{
            "Market": "KAR", "Genre": "GEC", "Channel": "CH", "Program": "P",
            "Days": "Mon-Sun", "Start Time": "18:00", "End Time": "22:00",
            "Net Rate 10Sec": 1000, "ACD": 20, "Spots": 140, "FCT": 2800,
        }]
        return upload_plan(rows)

    def test_weekly_grp_dispersion_biases_weeks(self):
        d = self._make_plan()
        gen = generate(d["plan_id"],
                       edits=[{"duration": 20, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-02-01",
                              "weekly_grp_dispersion": [10, 20, 30, 40]})
        wk = Counter(s["week"] for s in gen["schedule_rows"])
        assert wk[4] > wk[1], f"expected wk4>wk1 dispersion, got {dict(wk)}"
        # invariant
        assert sum(wk.values()) == sum(e["final_spots"] for e in gen["edit_rows"])

    def test_weekend_boost_puts_more_on_sat_sun(self):
        d = self._make_plan()
        gen = generate(d["plan_id"],
                       edits=[{"duration": 20, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-01-18",
                              "weekend_boost": 3.0})
        by_dow = Counter(s["day"] for s in gen["schedule_rows"])
        weekday_avg = sum(by_dow[d] for d in ("Mon", "Tue", "Wed", "Thu", "Fri")) / 5
        weekend_avg = (by_dow["Sat"] + by_dow["Sun"]) / 2
        assert weekend_avg > weekday_avg, f"weekend_boost failed: {dict(by_dow)}"

    def test_reach_vs_frequency_frequency_mode(self):
        d = self._make_plan()
        gen = generate(d["plan_id"],
                       edits=[{"duration": 20, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-01-18",
                              "reach_vs_frequency": 1.0})
        # invariant still holds
        assert len(gen["schedule_rows"]) == sum(e["final_spots"] for e in gen["edit_rows"])

    def test_gec_planning_weeks_limits_gec_only(self):
        # 4 week campaign, but gec_planning_weeks=2 -> GEC rows in wk1-2 only
        d = self._make_plan()
        gen = generate(d["plan_id"],
                       edits=[{"duration": 20, "percentage": 100}],
                       prefs={"campaign_start": "2026-01-05", "campaign_end": "2026-02-01",
                              "gec_planning_weeks": 2})
        wks = {s["week"] for s in gen["schedule_rows"]}
        assert wks.issubset({1, 2}), f"GEC leaked past wk2: {wks}"
        # invariant
        assert len(gen["schedule_rows"]) == sum(e["final_spots"] for e in gen["edit_rows"])
