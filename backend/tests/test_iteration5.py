"""Iteration 5 targeted tests: channel-order preservation, edit-order desc, regression."""
import os
import io
import openpyxl
import pytest
import requests

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://spot-scheduling-hub.preview.emergentagent.com",
).rstrip("/")
SAMPLE = "/tmp/samples/input.xlsx"


@pytest.fixture(scope="module")
def uploaded():
    with open(SAMPLE, "rb") as fh:
        r = requests.post(
            f"{BASE_URL}/api/plans/upload",
            files={"file": ("input.xlsx", fh,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="module")
def input_channel_order(uploaded):
    seen = []
    for row in uploaded["rows"]:
        ch = row.get("channel")
        if ch and ch not in seen:
            seen.append(ch)
    return seen


@pytest.fixture(scope="module")
def generated_wb(uploaded):
    pid = uploaded["plan_id"]
    payload = {
        # Deliberately out-of-order to prove sorting works
        "edits": [
            {"duration": 10, "percentage": 10},
            {"duration": 45, "percentage": 20},
            {"duration": 30, "percentage": 50},
            {"duration": 20, "percentage": 20},
        ],
        "prefs": {
            "campaign_start": "2026-06-29",
            "campaign_end": "2026-08-16",
            "campaign_weeks": 7,
            "spot_frequency_minutes": 30,
            "gec_genres": ["GEC"],
            "weekly_grp_dispersion": [15, 15, 15, 15, 15, 15, 10],
            "blackout_days": [],
            "blackout_dates": [],
            "daypart_weights": [],
        },
        "row_overrides": [],
    }
    r = requests.post(f"{BASE_URL}/api/plans/{pid}/generate", json=payload)
    assert r.status_code == 200, r.text
    result_id = r.json()["result_id"]
    r2 = requests.get(f"{BASE_URL}/api/results/{result_id}/download")
    assert r2.status_code == 200
    return openpyxl.load_workbook(io.BytesIO(r2.content), data_only=True)


def _find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and any(str(c or "").strip().lower() == "channel" for c in row):
            return i, list(row)
    return None, None


# ------------- CHANNEL ORDER -------------
def test_channel_order_matches_input(generated_wb, input_channel_order):
    ws = generated_wb["Schedule Sheet"]
    hdr_row, _ = _find_header_row(ws)
    assert hdr_row, "header row not found"

    seen = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        ch = row[3]
        if ch and isinstance(ch, str) and ch.strip():
            if any(k in ch.lower() for k in ("total", "grand", "subtotal")):
                continue
            if ch.strip() not in seen:
                seen.append(ch.strip())

    assert seen, "no channels emitted"
    # First data channel should equal first input channel
    assert seen[0] == input_channel_order[0], (
        f"first output channel {seen[0]!r} != first input channel "
        f"{input_channel_order[0]!r}"
    )
    # Order of common channels must match input order
    common_in_input = [c for c in input_channel_order if c in seen]
    common_in_output = [c for c in seen if c in common_in_input]
    assert common_in_output == common_in_input, (
        f"channel order mismatch:\ninput: {common_in_input}\noutput: {common_in_output}"
    )


# ------------- EDIT ORDER DESC -------------
def test_edit_durations_descending(generated_wb):
    ws = generated_wb["Schedule Sheet"]
    hdr_row, _ = _find_header_row(ws)
    # col 4 = Program (index 4), col 40 = Edit, col 13 = No (row id)
    # Collect all edit durations in order (skipping subtotal rows).
    # Then chunk by "new block starts whenever duration goes UP" (i.e. dur > prev).
    all_durs = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        prog = row[4]
        ed = row[40]
        if prog and isinstance(prog, str) and any(
            k in prog.lower() for k in ("total", "grand", "subtotal")
        ):
            all_durs.append(None)  # break marker
            continue
        if ed is None:
            continue
        try:
            all_durs.append(int(ed))
        except (ValueError, TypeError):
            continue

    # Chunk into monotonically-non-increasing blocks
    blocks = []
    cur = []
    for d in all_durs:
        if d is None:
            if cur:
                blocks.append(cur)
                cur = []
            continue
        if cur and d > cur[-1]:
            blocks.append(cur)
            cur = [d]
        else:
            cur.append(d)
    if cur:
        blocks.append(cur)

    checked = 0
    bad = []
    for durs in blocks:
        if len(durs) >= 2:
            if durs != sorted(durs, reverse=True):
                bad.append(durs)
            checked += 1
    assert checked > 0, "no multi-edit groups found"
    assert not bad, f"edit ordering violations: {bad[:5]}"


# ------------- REGRESSION: subtotals + weekly sections -------------
def test_regression_subtotals_and_weekly(generated_wb):
    ws = generated_wb["Schedule Sheet"]
    # collect strings from all cells
    strings = []
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if isinstance(c, str):
                strings.append(c.lower())
    all_text = " | ".join(strings)
    assert "grand total" in all_text, "Missing 'Grand Total'"
    # weekly sections: Wk 1..Wk 7 headers
    assert "wk 1" in all_text and "wk 7" in all_text, "Weekly section headers missing"
    # 3 weekly sections: check that 'wk 1' appears at least 3 times
    assert strings.count("wk 1") >= 3, (
        f"expected >=3 'Wk 1' headers (3 weekly sections), got {strings.count('wk 1')}"
    )
