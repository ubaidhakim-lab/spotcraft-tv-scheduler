"""Backend tests for ACD Plan Builder - Iteration 4.

Tests the new features:
1. Upload filters in-file subtotal rows -> 46 real rows (not 58)
2. campaign_end + derived weeks
3. blackout_dates ISO exclusion
4. schedule weeks start from campaign_start (no Monday alignment)
5. downloaded xlsx has subtotals per Channel/Genre/Market/GRAND + 3 weekly sections
"""
import io
import math
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL", "https://spot-scheduling-hub.preview.emergentagent.com"
).rstrip("/")
INPUT_XLSX = "/tmp/samples/input.xlsx"


@pytest.fixture(scope="module")
def uploaded_plan():
    with open(INPUT_XLSX, "rb") as f:
        files = {"file": ("input.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/plans/upload", files=files, timeout=60)
    assert r.status_code == 200, r.text
    return r.json()


class TestUploadSubtotalFilter:
    def test_row_count_is_46_real_programs(self, uploaded_plan):
        # After filtering rows where market/genre/channel/program contains 'Total'
        assert uploaded_plan["row_count"] == 46, f"expected 46, got {uploaded_plan['row_count']}"

    def test_no_total_rows_leaked(self, uploaded_plan):
        for row in uploaded_plan["rows"]:
            for k in ("market", "genre", "channel", "program"):
                v = str(row.get(k, "") or "").lower()
                assert "total" not in v, f"leaked total row: {row}"


@pytest.fixture(scope="module")
def generated_full(uploaded_plan):
    plan_id = uploaded_plan["plan_id"]
    payload = {
        "edits": [
            {"duration": 45, "percentage": 40},
            {"duration": 30, "percentage": 30},
            {"duration": 20, "percentage": 20},
            {"duration": 10, "percentage": 10},
        ],
        "row_overrides": [],
        "prefs": {
            "campaign_start": "2026-06-29",
            "campaign_end": "2026-08-16",
            "spot_frequency_minutes": 30,
            "gec_genres": ["GEC"],
            "weekly_grp_dispersion": [],
            "blackout_days": [],
            "blackout_dates": ["2026-07-04", "2026-07-15"],
            "daypart_weights": [],
        },
    }
    r = requests.post(f"{BASE_URL}/api/plans/{plan_id}/generate", json=payload, timeout=180)
    assert r.status_code == 200, r.text
    return r.json()


class TestDerivedWeeks:
    def test_weeks_is_7(self, generated_full):
        # (Aug 16 - Jun 29 + 1) = 49 days = ceil(49/7)=7
        by_week = generated_full["summary"]["by_week"]
        assert len(by_week) == 7, f"weeks={len(by_week)}"


class TestBlackoutDates:
    def test_no_schedule_on_blackout_dates(self, generated_full):
        for r in generated_full["schedule_rows"]:
            assert r["date"] not in ("2026-07-04", "2026-07-15"), f"row on blackout: {r}"


class TestDownloadSubtotalsAndWeekly:
    @pytest.fixture(scope="class")
    def workbook(self, generated_full):
        rid = generated_full["result_id"]
        r = requests.get(f"{BASE_URL}/api/results/{rid}/download", timeout=60)
        assert r.status_code == 200
        return load_workbook(io.BytesIO(r.content), data_only=True)

    def test_schedule_sheet_exists(self, workbook):
        assert "Schedule Sheet" in workbook.sheetnames

    def test_first_daily_column_is_campaign_start(self, workbook):
        ws = workbook["Schedule Sheet"]
        # Find the header row & find a cell containing '2026-06-29' or 'Jun 29'
        found = False
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            joined = " ".join(str(c) for c in row if c is not None)
            if "2026-06-29" in joined or "Jun 29" in joined or "29-Jun" in joined or "Jun-29" in joined:
                found = True
                break
        assert found, "campaign_start (2026-06-29 / Jun 29) not found in header rows"

    def test_subtotals_present(self, workbook):
        ws = workbook["Schedule Sheet"]
        text_cells = []
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c is not None:
                    text_cells.append(str(c))
        blob = " | ".join(text_cells)
        # Expect various subtotal markers
        assert "GRAND TOTAL" in blob, "GRAND TOTAL row missing"
        # channel/genre/market total
        total_markers = [s for s in text_cells if "Total" in s and s != "GRAND TOTAL"]
        assert len(total_markers) >= 3, f"expected channel/genre/market subtotal rows, got {total_markers}"

    def test_three_weekly_sections(self, workbook):
        ws = workbook["Schedule Sheet"]
        blob = " ".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None)
        assert "Weekly Spots" in blob, "Weekly Spots banner missing"
        assert "Weekly Spot Dispersion" in blob, "Weekly Spot Dispersion banner missing"
        assert "Weekly GRP" in blob, "Weekly GRP banner missing"
