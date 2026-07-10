"""Backend tests for ACD Plan Builder - Iteration 2.
Uses real input/output sample files at /tmp/samples/.
"""
import os
import io
import pytest
import requests

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://spot-scheduling-hub.preview.emergentagent.com').rstrip('/')
INPUT_XLSX = "/tmp/samples/input.xlsx"
OUTPUT_XLSX = "/tmp/samples/output.xlsx"


@pytest.fixture(scope="module")
def uploaded_plan():
    with open(INPUT_XLSX, "rb") as f:
        files = {"file": ("input.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/plans/upload", files=files, timeout=60)
    assert r.status_code == 200, r.text
    return r.json()


# ---------------- Upload ----------------
class TestUpload:
    def test_upload_real_input(self, uploaded_plan):
        data = uploaded_plan
        assert "plan_id" in data
        assert data["row_count"] >= 50, f"Expected >=50 rows, got {data['row_count']}"
        assert data["summary"]["total_spots"] >= 1000, f"total_spots={data['summary']['total_spots']}"
        # metadata should include Client/Campaign
        md = data.get("metadata", {})
        md_keys_lower = {k.lower(): v for k, v in md.items()}
        assert any("client" in k for k in md_keys_lower), f"metadata keys: {list(md.keys())}"

    def test_spots_canonical_not_from_cal_spts_pd(self, uploaded_plan):
        # spots for each row must be numeric and > 0 mostly
        rows = uploaded_plan["rows"]
        nonzero = sum(1 for r in rows if (r.get("spots") or 0) > 0)
        assert nonzero >= 30


# ---------------- Learn sample ----------------
class TestLearnSample:
    def test_learn_output(self):
        with open(OUTPUT_XLSX, "rb") as f:
            files = {"file": ("output.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = requests.post(f"{BASE_URL}/api/plans/learn-sample", files=files, timeout=60)
        assert r.status_code == 200, r.text
        data = r.json()
        assert "edits" in data
        assert isinstance(data["edits"], list)
        if data["edits"]:
            total = sum(e["percentage"] for e in data["edits"])
            assert 95 <= total <= 105, f"edits sum={total}"
        assert "weekly_grp_dispersion" in data


# ---------------- Generate ----------------
@pytest.fixture(scope="module")
def generated(uploaded_plan):
    plan_id = uploaded_plan["plan_id"]
    payload = {
        "edits": [
            {"duration": 45, "percentage": 10},
            {"duration": 30, "percentage": 40},
            {"duration": 20, "percentage": 30},
            {"duration": 10, "percentage": 20},
        ],
        "row_overrides": [],
        "prefs": {
            "campaign_start": "2025-01-06",
            "campaign_weeks": 7,
            "spot_frequency_minutes": 30,
            "gec_genres": ["GEC"],
            "weekly_grp_dispersion": [],
            "blackout_days": [],
            "daypart_weights": []
        }
    }
    r = requests.post(f"{BASE_URL}/api/plans/{plan_id}/generate", json=payload, timeout=120)
    assert r.status_code == 200, r.text
    return r.json()


class TestGenerate:
    def test_edit_rows_count(self, generated, uploaded_plan):
        expected = uploaded_plan["row_count"] * 4
        assert len(generated["edit_rows"]) == expected, f"got {len(generated['edit_rows'])} vs expected {expected}"

    def test_schedule_rows_thousands(self, generated):
        assert len(generated["schedule_rows"]) >= 1000

    def test_summary_has_by_week(self, generated):
        assert len(generated["summary"]["by_week"]) == 7


# ---------------- Download ----------------
class TestDownload:
    def test_download_xlsx(self, generated):
        rid = generated["result_id"]
        r = requests.get(f"{BASE_URL}/api/results/{rid}/download", timeout=60)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), data_only=True)
        assert "Schedule Sheet" in wb.sheetnames
        ws = wb["Schedule Sheet"]
        assert ws.max_row > 10


# ---------------- Row override ----------------
class TestRowOverride:
    def test_override_row0(self, uploaded_plan):
        plan_id = uploaded_plan["plan_id"]
        payload = {
            "edits": [
                {"duration": 45, "percentage": 10},
                {"duration": 30, "percentage": 40},
                {"duration": 20, "percentage": 30},
                {"duration": 10, "percentage": 20},
            ],
            "row_overrides": [{"row_id": 0, "edits": [{"duration": 30, "percentage": 100}]}],
            "prefs": {
                "campaign_start": "2025-01-06",
                "campaign_weeks": 4,
                "spot_frequency_minutes": 30,
                "gec_genres": ["GEC"],
                "weekly_grp_dispersion": [],
                "blackout_days": [],
                "daypart_weights": []
            }
        }
        r = requests.post(f"{BASE_URL}/api/plans/{plan_id}/generate", json=payload, timeout=120)
        assert r.status_code == 200, r.text
        data = r.json()
        row0_edits = [e for e in data["edit_rows"] if e["_row_id"] == 0]
        assert len(row0_edits) == 1
        assert row0_edits[0]["edit_duration"] == 30
        assert row0_edits[0]["edit_pct"] == 100


# ---------------- Daypart weights ----------------
class TestDaypartWeights:
    def test_daypart_weights_no_crash(self, uploaded_plan):
        plan_id = uploaded_plan["plan_id"]
        payload = {
            "edits": [{"duration": 30, "percentage": 100}],
            "row_overrides": [],
            "prefs": {
                "campaign_start": "2025-01-06",
                "campaign_weeks": 2,
                "spot_frequency_minutes": 30,
                "gec_genres": ["GEC"],
                "weekly_grp_dispersion": [],
                "blackout_days": [],
                "daypart_weights": [{"daypart": "Prime Time", "weight": 3}]
            }
        }
        r = requests.post(f"{BASE_URL}/api/plans/{plan_id}/generate", json=payload, timeout=120)
        assert r.status_code == 200, r.text
        assert len(r.json()["schedule_rows"]) > 0


# ---------------- Sessions ----------------
class TestSessions:
    def test_sessions_crud(self, uploaded_plan):
        plan_id = uploaded_plan["plan_id"]
        payload = {
            "name": "TEST_session_1",
            "plan_id": plan_id,
            "edits": [{"duration": 30, "percentage": 100}],
            "row_overrides": [],
            "prefs": {
                "campaign_start": "2025-01-06",
                "campaign_weeks": 4,
                "spot_frequency_minutes": 30,
                "gec_genres": ["GEC"],
                "weekly_grp_dispersion": [],
                "blackout_days": [],
                "daypart_weights": []
            }
        }
        r = requests.post(f"{BASE_URL}/api/sessions", json=payload, timeout=30)
        assert r.status_code == 200, r.text
        sid = r.json()["id"]

        r2 = requests.get(f"{BASE_URL}/api/sessions", timeout=30)
        assert r2.status_code == 200
        assert any(s["id"] == sid for s in r2.json())

        r3 = requests.get(f"{BASE_URL}/api/sessions/{sid}", timeout=30)
        assert r3.status_code == 200
        assert r3.json()["name"] == "TEST_session_1"

        r4 = requests.delete(f"{BASE_URL}/api/sessions/{sid}", timeout=30)
        assert r4.status_code == 200

        r5 = requests.get(f"{BASE_URL}/api/sessions/{sid}", timeout=30)
        assert r5.status_code == 404
