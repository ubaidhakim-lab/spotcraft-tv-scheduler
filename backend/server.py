from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
import math
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta, date, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# -------------------- Models --------------------

class EditConfig(BaseModel):
    duration: int
    percentage: float

class RowOverride(BaseModel):
    row_id: int
    edits: List[EditConfig]

class DaypartWeight(BaseModel):
    daypart: str
    weight: float

class SchedulingPrefs(BaseModel):
    campaign_start: str
    campaign_end: Optional[str] = None
    campaign_weeks: int = 6
    spot_frequency_minutes: int = 30
    movies_frequency_minutes: int = 60  # 1 spot every 60 min on movie channels
    movies_genres: List[str] = ["MOV", "Movies", "Movie"]
    gec_genres: List[str] = ["GEC"]
    gec_planning_weeks: Optional[int] = None
    weekly_grp_dispersion: List[float] = []
    blackout_days: List[str] = []
    blackout_dates: List[str] = []
    daypart_weights: List[DaypartWeight] = []
    weekend_boost: float = 1.0  # multiplier on Sat/Sun; 1 = neutral, >1 = more, <1 = less
    reach_vs_frequency: float = 0.5  # 0 = max reach (spread across days), 1 = max frequency (concentrate)

class GenerateRequest(BaseModel):
    edits: List[EditConfig]
    row_overrides: List[RowOverride] = []
    prefs: SchedulingPrefs

class SessionSave(BaseModel):
    name: str
    plan_id: str
    edits: List[EditConfig]
    row_overrides: List[RowOverride] = []
    prefs: SchedulingPrefs

# -------------------- Helpers --------------------

DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

DAY_ALIASES = {
    "mon": "Mon", "monday": "Mon", "m": "Mon",
    "tue": "Tue", "tues": "Tue", "tuesday": "Tue",
    "wed": "Wed", "wednesday": "Wed", "w": "Wed",
    "thu": "Thu", "thur": "Thu", "thurs": "Thu", "thursday": "Thu", "th": "Thu",
    "fri": "Fri", "friday": "Fri", "f": "Fri",
    "sat": "Sat", "saturday": "Sat",
    "sun": "Sun", "sunday": "Sun", "su": "Sun",
}

CANONICAL_ALIASES = {
    "market": ["market", "markets"],
    "genre": ["genre"],
    "channel": ["channel"],
    "program": ["program", "programme", "show"],
    "days": ["days", "day"],
    "start_time": ["start_time", "starttime", "start"],
    "end_time": ["end_time", "endtime", "end"],
    "net_rate_10s": ["net_rate_10sec", "nett_rate_10sec", "netrate10sec", "nett_rate_10s", "net_rate", "rate"],
    "acd": ["acd", "average_commercial_duration"],
    "spots": ["spots", "final_spots", "my_disp_spots"],
    "fct": ["fct", "final_fct", "planned_fct_per_tb"],
    "outlay": ["net_outlay", "outlay", "cost"],
    "log_tvr": ["log_tvr", "logtvr", "tvr"],
    "grp": ["grp", "grps"],
    "ngrp": ["ngrp"],
    "cprp": ["cprp"],
    "direct_matrix": ["direct_matrix", "direct"],
    "index_value": ["index_value", "index"],
}

def normalize_key(s: str) -> str:
    return str(s).strip().lower().replace(" ", "_").replace("/", "_").replace(".", "").replace("-", "_").replace("__", "_")

def to_serializable(v):
    if v is None:
        return None
    if isinstance(v, (datetime,)):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        h = total // 3600
        m = (total % 3600) // 60
        s = total % 60
        return f"{h:02d}:{m:02d}:{s:02d}"
    if isinstance(v, float) and math.isnan(v):
        return None
    return v

def parse_days(days_str) -> List[str]:
    if days_str is None:
        return DAY_ORDER[:]
    s = str(days_str).strip().lower()
    if not s or s == "nan":
        return DAY_ORDER[:]
    if s in ("daily", "all", "all days", "everyday", "mon-sun"):
        return DAY_ORDER[:]
    if s in ("mon-fri", "mon-friday", "weekday", "weekdays", "mtwtf"):
        return DAY_ORDER[:5]
    if s in ("mon-sat", "mon-saturday"):
        return DAY_ORDER[:6]
    if s in ("sat-sun", "weekend", "weekends", "sat,sun", "sat-sunday"):
        return DAY_ORDER[5:]
    if s in ("sun-sat",):
        return DAY_ORDER[:]
    # range like tue-fri
    if "-" in s and "," not in s:
        parts = s.split("-")
        if len(parts) == 2 and parts[0].strip() in DAY_ALIASES and parts[1].strip() in DAY_ALIASES:
            a = DAY_ALIASES[parts[0].strip()]
            b = DAY_ALIASES[parts[1].strip()]
            ia, ib = DAY_ORDER.index(a), DAY_ORDER.index(b)
            if ia <= ib:
                return DAY_ORDER[ia:ib + 1]
            return DAY_ORDER[ia:] + DAY_ORDER[:ib + 1]
    for sep in [",", "/", "+", "|", "&"]:
        s = s.replace(sep, " ")
    tokens = [t for t in s.split() if t]
    result = []
    for t in tokens:
        if t in DAY_ALIASES and DAY_ALIASES[t] not in result:
            result.append(DAY_ALIASES[t])
    return result or DAY_ORDER[:]

def parse_time(t) -> Optional[timedelta]:
    if t is None:
        return None
    if isinstance(t, time):
        return timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
    if isinstance(t, datetime):
        return timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
    if isinstance(t, timedelta):
        return t
    s = str(t).strip()
    if not s or s.lower() == "nan":
        return None
    s = s.replace(".", ":")
    for fmt in ["%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p", "%I:%M%p", "%I %p"]:
        try:
            dt = datetime.strptime(s.upper(), fmt)
            return timedelta(hours=dt.hour, minutes=dt.minute, seconds=dt.second)
        except ValueError:
            continue
    try:
        val = float(s)
        if val < 1:
            return timedelta(seconds=int(val * 86400))
    except Exception:
        pass
    return None

def format_time(td: timedelta) -> str:
    total = int(td.total_seconds()) % 86400
    h = total // 3600
    m = (total % 3600) // 60
    return f"{h:02d}:{m:02d}"

def safe_num(v, default=0.0):
    try:
        if v is None:
            return default
        if isinstance(v, bool):
            return float(v)
        if isinstance(v, str):
            s = v.strip().replace(",", "")
            if not s or s.startswith("#"):
                return default
            return float(s)
        f = float(v)
        if math.isnan(f):
            return default
        return f
    except Exception:
        return default

def daypart_of(td: timedelta) -> str:
    h = int(td.total_seconds()) // 3600
    if 6 <= h < 9:
        return "Morning"
    if 9 <= h < 12:
        return "Late Morning"
    if 12 <= h < 15:
        return "Afternoon"
    if 15 <= h < 18:
        return "Late Afternoon"
    if 18 <= h < 21:
        return "Prime Time"
    if 21 <= h < 24:
        return "Late Prime"
    return "Overnight"

# -------------------- Sheet parsing --------------------

def find_header_row(ws) -> int:
    """Find the row containing column headers. Look for a row that contains 'Program' or ('Genre' AND 'Channel')."""
    max_scan = min(ws.max_row, 25)
    for r in range(1, max_scan + 1):
        vals = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[r]]
        has_program = any("program" == v for v in vals)
        has_channel = any("channel" == v for v in vals)
        has_genre = any("genre" == v for v in vals)
        if has_program and (has_channel or has_genre):
            return r
    return 1  # fallback

def resolve_canonical(header_norm: str) -> Optional[str]:
    for canonical, aliases in CANONICAL_ALIASES.items():
        if header_norm in aliases:
            return canonical
    return None

def read_workbook(content: bytes):
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = find_header_row(ws)

    # Read metadata (any rows above header where col2 has a "key" and col3 has value)
    metadata = {}
    for r in range(1, header_row):
        row = ws[r]
        for j in range(0, min(6, len(row))):
            key = row[j].value
            if key and isinstance(key, str) and j + 1 < len(row):
                val = row[j + 1].value
                if val is not None and str(key).strip().lower() in ("client", "brand", "campaign", "period", "campaign period", "tg", "markets", "market"):
                    metadata[str(key).strip()] = to_serializable(val)

    # Read headers
    raw_headers = [c.value for c in ws[header_row]]
    headers = []
    canonical_of = {}
    seen_canonical = set()
    for j, h in enumerate(raw_headers):
        if h is None or str(h).strip() == "":
            headers.append(f"_col{j+1}")
            continue
        base = str(h).strip()
        norm = normalize_key(base)
        canonical = resolve_canonical(norm)
        # If the canonical was already claimed, keep as extra column with suffix
        if canonical and canonical not in seen_canonical:
            seen_canonical.add(canonical)
            canonical_of[base] = canonical
            headers.append(base)
        else:
            # unique-ify duplicate names
            key = base
            k = 2
            while key in headers:
                key = f"{base} ({k})"
                k += 1
            headers.append(key)

    # Read data rows
    raw_rows = []
    parsed_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = ws[r]
        vals = [to_serializable(c.value) for c in row]
        # skip completely empty rows
        if all(v is None or v == "" for v in vals):
            continue
        # skip subtotal/summary rows: if col1 or col-first non-null contains "Total"
        first_text = next((str(v) for v in vals if v is not None), "")
        if "total" in first_text.lower() and not first_text.lower().startswith("total") and "kar" not in first_text.lower():
            # Heuristic: row where any column value contains 'Total' alone
            joined = " ".join(str(v) for v in vals if v is not None)
            if joined.strip().endswith("Total") or " Total " in joined:
                continue
        raw = {}
        for j, v in enumerate(vals):
            if j < len(headers):
                raw[headers[j]] = v
        raw["_row_id"] = len(raw_rows)
        raw_rows.append(raw)

        # Canonical parsed row
        p = {"_row_id": raw["_row_id"]}
        for base, canonical in canonical_of.items():
            v = raw.get(base)
            if canonical in ("net_rate_10s", "acd", "spots", "fct", "outlay", "log_tvr", "grp", "ngrp", "cprp", "index_value"):
                p[canonical] = safe_num(v)
            else:
                p[canonical] = v
        parsed_rows.append(p)

    # Detect skipped subtotals a second way: rows where core fields are all missing
    filtered = []
    filtered_raw = []
    for i, pr in enumerate(parsed_rows):
        # require at least channel or program
        if not pr.get("channel") and not pr.get("program"):
            continue
        # skip subtotal rows: Program, Channel, or Genre value contains "Total" (case-insensitive)
        subtotal_hit = False
        for key in ("program", "channel", "genre", "market"):
            v = pr.get(key)
            if v and "total" in str(v).lower():
                subtotal_hit = True
                break
        if subtotal_hit:
            continue
        filtered.append(pr)
        filtered_raw.append(raw_rows[i])

    # renumber row_id
    for i, r in enumerate(filtered):
        r["_row_id"] = i
    for i, r in enumerate(filtered_raw):
        r["_row_id"] = i

    return {
        "metadata": metadata,
        "header_row": header_row,
        "columns": headers,
        "raw_rows": filtered_raw,
        "parsed_rows": filtered,
    }

# -------------------- Endpoints --------------------

@api_router.get("/")
async def root():
    return {"message": "ACD Plan Builder API"}

@api_router.post("/plans/upload")
async def upload_plan(file: UploadFile = File(...)):
    content = await file.read()
    try:
        parsed = read_workbook(content)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}") from e

    plan_id = str(uuid.uuid4())
    doc = {
        "id": plan_id,
        "filename": file.filename,
        "metadata": parsed["metadata"],
        "header_row": parsed["header_row"],
        "columns": parsed["columns"],
        "raw_rows": parsed["raw_rows"],
        "parsed_rows": parsed["parsed_rows"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.plans.insert_one(doc)

    rows = parsed["parsed_rows"]
    total_fct = sum(safe_num(r.get("fct")) for r in rows)
    total_spots = sum(safe_num(r.get("spots")) for r in rows)
    total_grp = sum(safe_num(r.get("grp")) for r in rows)
    total_outlay = sum(safe_num(r.get("outlay")) for r in rows)

    return {
        "plan_id": plan_id,
        "filename": file.filename,
        "metadata": parsed["metadata"],
        "columns": parsed["columns"],
        "row_count": len(rows),
        "rows": rows[:200],  # canonical preview
        "raw_preview": parsed["raw_rows"][:20],
        "summary": {
            "total_fct": total_fct,
            "total_spots": total_spots,
            "total_grp": total_grp,
            "total_outlay": total_outlay,
        },
    }

@api_router.post("/plans/learn-sample")
async def learn_sample(file: UploadFile = File(...)):
    """Learn edit dispersion + weekly dispersion from a past output schedule."""
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row = find_header_row(ws)
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[header_row]]
        headers_lower = [h.lower() for h in headers]

        # Find Edit column & Final Spots column
        edit_col = None
        spots_col = None
        for j, h in enumerate(headers_lower):
            if h == "edit":
                edit_col = j
            if h == "final spots" or h == "spots" or h == "cal spts pd":
                if spots_col is None:
                    spots_col = j

        # Sum spots by edit duration
        edit_spots: Dict[int, float] = {}
        # Also detect week columns (WK 1, WK 2 ... in header row 10 area OR find row with "Wk 1" etc)
        week_totals: Dict[int, float] = {}
        # Search for weekly summary columns
        week_cols = []
        for j, h in enumerate(headers_lower):
            if h.startswith("wk ") and h[3:].strip().isdigit():
                week_cols.append((int(h[3:].strip()), j))

        for r in range(header_row + 1, ws.max_row + 1):
            row = ws[r]
            if edit_col is not None and spots_col is not None:
                e = safe_num(row[edit_col].value)
                s = safe_num(row[spots_col].value)
                if e > 0 and s > 0:
                    edit_spots[int(round(e))] = edit_spots.get(int(round(e)), 0) + s
            for wk_num, jcol in week_cols:
                v = safe_num(row[jcol].value)
                week_totals[wk_num] = week_totals.get(wk_num, 0) + v

        total_spots = sum(edit_spots.values())
        edits = []
        if total_spots > 0:
            for dur, sp in sorted(edit_spots.items(), key=lambda x: -x[1]):
                edits.append({"duration": dur, "percentage": round(sp * 100.0 / total_spots, 2)})

        wk_disp = []
        if week_totals:
            wk_total = sum(week_totals.values())
            if wk_total > 0:
                for w in sorted(week_totals.keys()):
                    wk_disp.append(round(week_totals[w] * 100.0 / wk_total, 2))

        return {
            "edits": edits,
            "weekly_grp_dispersion": wk_disp,
            "campaign_weeks": len(wk_disp) if wk_disp else 0,
            "source_filename": file.filename,
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to learn from sample: {e}")

@api_router.get("/plans/{plan_id}")
async def get_plan(plan_id: str):
    doc = await db.plans.find_one({"id": plan_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Plan not found")
    return doc

def apply_daypart_weights(slot_times: List[timedelta], weights: Dict[str, float]) -> List[float]:
    if not weights:
        return [1.0] * len(slot_times)
    out = []
    for t in slot_times:
        dp = daypart_of(t)
        w = weights.get(dp, 1.0)
        out.append(max(0.01, w))
    total = sum(out)
    if total <= 0:
        return [1.0] * len(slot_times)
    return [x / total * len(slot_times) for x in out]

def allocate_spots_daily(
    days_list: List[str],
    slot_times: List[timedelta],
    slot_weights: List[float],
    count: int,
    week_start: date,
    blackout_dates: set = None,
    campaign_end: Optional[date] = None,
    week_index: int = 0,
    weekend_boost: float = 1.0,
    reach_vs_frequency: float = 0.5,
    used_slots: Optional[set] = None,
):
    """Return list of (day_name, date, timedelta) for scheduled spots.

    Never places a spot on a day-of-week outside days_list, past campaign_end,
    on a blackout_date, or on a (date, time) already present in used_slots.
    """
    blackout_dates = blackout_dates or set()
    used_slots = used_slots if used_slots is not None else set()
    # Build up to 7 dates of this week; skip dates beyond campaign_end
    day_dates: Dict[str, date] = {}
    for i in range(7):
        d = week_start + timedelta(days=i)
        if campaign_end and d > campaign_end:
            break
        dow = DAY_ORDER[d.weekday()]
        if dow in days_list and d not in blackout_dates and dow not in day_dates:
            day_dates[dow] = d
    if not day_dates:
        return []
    # Keep only days_list days that actually have a date this week
    active_day_names = [d for d in days_list if d in day_dates]
    if not active_day_names:
        return []
    # Rotate start-of-week to balance across all valid days over the campaign
    if len(active_day_names) > 1:
        off = week_index % len(active_day_names)
        active_day_names = active_day_names[off:] + active_day_names[:off]
    # Apply weekend boost to slot_weights per day (Sat/Sun weighted differently)
    weekend_factor: Dict[str, float] = {}
    for d in active_day_names:
        weekend_factor[d] = weekend_boost if d in ("Sat", "Sun") else 1.0
    # Build slot pool per day; each entry is a (weight, time, date) tuple.
    by_day: Dict[str, List] = {d: [] for d in active_day_names}
    for d in active_day_names:
        wf = weekend_factor[d]
        for i, t in enumerate(slot_times):
            by_day[d].append((slot_weights[i] * wf, t, day_dates[d]))
        by_day[d].sort(key=lambda x: (-x[0], x[1]))
    allocated = []
    used: set = set(used_slots)

    def _slot_key(dt, t):
        return (dt, format_time(t))

    # Compute per-day target quotas proportional to weekend_factor weights.
    # This is how weekend_boost actually shifts spot distribution across days.
    day_weight = {d: weekend_factor[d] for d in active_day_names}
    total_w = sum(day_weight.values())
    if total_w <= 0:
        total_w = float(len(active_day_names))
    raw_quota = {d: count * day_weight[d] / total_w for d in active_day_names}
    day_quota = {d: int(raw_quota[d]) for d in active_day_names}
    remainder = count - sum(day_quota.values())
    frac_order = sorted(
        active_day_names, key=lambda d: (-(raw_quota[d] - day_quota[d]), -day_weight[d])
    )
    for k in range(remainder):
        day_quota[frac_order[k % len(frac_order)]] += 1

    def _try_pick_from_day(d):
        """Pop the next free slot for day d, honoring `used`. Return (t, dt) or None."""
        while by_day[d]:
            w, t, dt = by_day[d].pop(0)
            if _slot_key(dt, t) not in used:
                return (t, dt)
        return None

    if reach_vs_frequency <= 0.5:
        # Reach mode: round-robin across days, but each day capped at its quota
        remaining = dict(day_quota)
        # Weighted round-robin: visit each day proportional to remaining quota
        while sum(remaining.values()) > 0:
            picked_any = False
            for d in active_day_names:
                if remaining[d] <= 0 or not by_day[d]:
                    continue
                pick = _try_pick_from_day(d)
                if pick is None:
                    remaining[d] = 0
                    continue
                t, dt = pick
                allocated.append((d, dt, t))
                used.add(_slot_key(dt, t))
                remaining[d] -= 1
                picked_any = True
            if not picked_any:
                break
    else:
        # Frequency mode: fill each day (heaviest-weight first) up to its quota
        order = sorted(active_day_names, key=lambda d: -day_weight[d])
        for d in order:
            need = day_quota[d]
            while need > 0:
                pick = _try_pick_from_day(d)
                if pick is None:
                    break
                t, dt = pick
                allocated.append((d, dt, t))
                used.add(_slot_key(dt, t))
                need -= 1

    # If quotas dropped spots (day ran out of slots but had remaining quota),
    # rebalance the leftover to any other day with slots.
    shortfall = count - len(allocated)
    if shortfall > 0:
        pool = [d for d in active_day_names if by_day[d]]
        while shortfall > 0 and pool:
            progressed = False
            for d in list(pool):
                pick = _try_pick_from_day(d)
                if pick is None:
                    pool.remove(d)
                    continue
                t, dt = pick
                allocated.append((d, dt, t))
                used.add(_slot_key(dt, t))
                shortfall -= 1
                progressed = True
                if shortfall == 0:
                    break
            if not progressed:
                break

    return allocated

@api_router.post("/plans/{plan_id}/generate")
async def generate_plan(plan_id: str, req: GenerateRequest):
    doc = await db.plans.find_one({"id": plan_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Plan not found")

    edits_global = req.edits
    total_pct = sum(e.percentage for e in edits_global)
    if abs(total_pct - 100) > 0.5:
        raise HTTPException(status_code=400, detail=f"Edit percentages must sum to 100 (got {total_pct})")

    # Row overrides
    override_map: Dict[int, List[EditConfig]] = {}
    for ov in req.row_overrides:
        s = sum(e.percentage for e in ov.edits)
        if abs(s - 100) > 0.5:
            raise HTTPException(status_code=400, detail=f"Row {ov.row_id} override percentages must sum to 100 (got {s})")
        override_map[ov.row_id] = ov.edits

    prefs = req.prefs
    try:
        camp_start = datetime.strptime(prefs.campaign_start, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid campaign_start (YYYY-MM-DD)") from None

    # Determine weeks: prefer explicit end date, else campaign_weeks
    if prefs.campaign_end:
        try:
            camp_end = datetime.strptime(prefs.campaign_end, "%Y-%m-%d").date()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid campaign_end (YYYY-MM-DD)") from None
        if camp_end < camp_start:
            raise HTTPException(status_code=400, detail="campaign_end must be on or after campaign_start")
        total_days = (camp_end - camp_start).days + 1
        weeks = max(1, math.ceil(total_days / 7))
    else:
        weeks = max(1, int(prefs.campaign_weeks))
        camp_end = camp_start + timedelta(days=weeks * 7 - 1)

    # Weeks start on campaign_start date (no Monday alignment)
    schedule_start = camp_start

    if prefs.weekly_grp_dispersion and len(prefs.weekly_grp_dispersion) == weeks:
        wk_disp = prefs.weekly_grp_dispersion[:]
    else:
        wk_disp = [100.0 / weeks] * weeks
    ws_sum = sum(wk_disp)
    wk_disp = [w * 100.0 / ws_sum if ws_sum else 0 for w in wk_disp]

    # Parse blackout dates (ISO YYYY-MM-DD)
    blackout_date_set = set()
    for ds in (prefs.blackout_dates or []):
        try:
            blackout_date_set.add(datetime.strptime(ds, "%Y-%m-%d").date())
        except Exception:
            continue

    daypart_weight_map = {d.daypart: d.weight for d in prefs.daypart_weights} if prefs.daypart_weights else {}

    # Blackout day-of-week names (e.g. "Sun") — kept alongside blackout_date_set (specific dates)
    blackout = set(prefs.blackout_days or [])

    def is_gec(g):
        return any(k.lower() in str(g or "").lower() for k in prefs.gec_genres)

    def is_movies(g):
        return any(k.lower() in str(g or "").lower() for k in (prefs.movies_genres or []))

    edit_rows: List[Dict[str, Any]] = []

    def compute_edit_capacity(row: Dict, edit_duration: int) -> int:
        """Max possible spots for this program-timeband-edit given the campaign window.

        capacity = eligible_days_per_week × slots_per_day_at_this_frequency × eligible_weeks
        minus dates in blackout that would have been active.
        """
        days_list = [d for d in parse_days(row.get("days")) if d not in blackout]
        if not days_list:
            return 0
        st = parse_time(row.get("start_time")) or timedelta(hours=6)
        et = parse_time(row.get("end_time")) or (st + timedelta(hours=1))
        if et <= st:
            et = st + timedelta(hours=1)
        freq = prefs.movies_frequency_minutes if is_movies(row.get("genre")) else prefs.spot_frequency_minutes
        step = timedelta(minutes=max(5, freq))
        slots_per_day = 0
        cur = st
        while cur < et:
            slots_per_day += 1
            cur += step
        if slots_per_day == 0:
            slots_per_day = 1
        eff_weeks = weeks
        if is_gec(row.get("genre")) and prefs.gec_planning_weeks:
            eff_weeks = min(weeks, prefs.gec_planning_weeks)
        # Estimate blackout impact — count blackout dates that fall on allowed days within eligible weeks
        blk_hits = 0
        for bd in blackout_date_set:
            dow = DAY_ORDER[bd.weekday()]
            if dow not in days_list:
                continue
            wk_idx = (bd - schedule_start).days // 7
            if 0 <= wk_idx < eff_weeks:
                blk_hits += 1
        capacity = len(days_list) * slots_per_day * eff_weeks - blk_hits * slots_per_day
        return max(0, capacity)

    for r in doc["parsed_rows"]:
        fct = safe_num(r.get("fct"))
        spots_total = safe_num(r.get("spots"))
        grp_total = safe_num(r.get("grp"))
        rate_10s = safe_num(r.get("net_rate_10s"))
        acd = safe_num(r.get("acd")) or 10

        # Determine actual FCT if missing: spots × ACD
        if fct <= 0 and spots_total > 0 and acd > 0:
            fct = spots_total * acd

        row_edits = override_map.get(r["_row_id"], edits_global)
        per_spot_grp = (grp_total / spots_total) if spots_total > 0 else 0

        # First pass: compute raw demand + capacity per edit
        demands = []
        capacities = []
        for e in row_edits:
            edit_fct = fct * (e.percentage / 100.0)
            raw = edit_fct / e.duration if e.duration > 0 else 0
            demands.append(int(round(raw)))
            capacities.append(compute_edit_capacity(r, e.duration))

        # Cap each edit at its capacity; try to rebalance any surplus into siblings with headroom
        capped = [min(demands[i], capacities[i]) for i in range(len(row_edits))]
        surplus = sum(demands) - sum(capped)
        if surplus > 0:
            # Distribute surplus to edits with headroom (in order of largest headroom first)
            while surplus > 0:
                headroom = [(capacities[i] - capped[i], i) for i in range(len(row_edits))]
                headroom = [(h, i) for h, i in headroom if h > 0]
                if not headroom:
                    break
                headroom.sort(reverse=True)
                added = False
                for h, i in headroom:
                    if surplus <= 0:
                        break
                    take = min(h, surplus)
                    capped[i] += take
                    surplus -= take
                    added = True
                if not added:
                    break

        # Second pass: build edit_rows using capped counts
        for idx, e in enumerate(row_edits):
            spots_int = capped[idx]
            final_fct = spots_int * e.duration
            net_outlay = final_fct * (rate_10s / 10.0)
            grp_share = spots_int * per_spot_grp

            edit_rows.append({
                "_row_id": r["_row_id"],
                "direct_matrix": r.get("direct_matrix"),
                "index_value": r.get("index_value") or (r["_row_id"] + 1),
                "market": r.get("market"),
                "genre": r.get("genre"),
                "channel": r.get("channel"),
                "program": r.get("program"),
                "days": r.get("days"),
                "start_time": r.get("start_time"),
                "end_time": r.get("end_time"),
                "net_rate_10s": rate_10s,
                "acd": acd,
                "edit_duration": e.duration,
                "edit_pct": e.percentage,
                "final_spots": spots_int,
                "final_fct": round(final_fct, 2),
                "net_outlay": round(net_outlay, 2),
                "grp": round(grp_share, 4),
                "log_tvr": safe_num(r.get("log_tvr")),
            })

    # -------- Day-wise schedule --------
    schedule_rows: List[Dict[str, Any]] = []

    # Global slot occupancy: per (channel, program, timeband, edit_duration)
    slot_occupancy: Dict[tuple, set] = {}

    for er in edit_rows:
        n_spots = er["final_spots"]
        if n_spots <= 0:
            continue
        days_list = [d for d in parse_days(er.get("days")) if d not in blackout]
        if not days_list:
            continue

        st = parse_time(er.get("start_time")) or timedelta(hours=6)
        et = parse_time(er.get("end_time")) or (st + timedelta(hours=1))
        if et <= st:
            et = st + timedelta(hours=1)

        # Movies channels use 60-min default spot frequency
        freq_min = prefs.movies_frequency_minutes if is_movies(er.get("genre")) else prefs.spot_frequency_minutes
        step = timedelta(minutes=max(5, freq_min))
        slot_times = []
        cur = st
        while cur < et:
            slot_times.append(cur)
            cur += step
        if not slot_times:
            slot_times = [st]
        slot_weights = apply_daypart_weights(slot_times, daypart_weight_map)

        eff_weeks = weeks
        if is_gec(er.get("genre")) and prefs.gec_planning_weeks:
            eff_weeks = min(weeks, prefs.gec_planning_weeks)

        base_disp = wk_disp[:eff_weeks]
        s = sum(base_disp)
        if s == 0:
            base_disp = [100.0 / eff_weeks] * eff_weeks
            s = 100.0
        norm_disp = [x * 100.0 / s for x in base_disp]
        raw = [n_spots * (p / 100.0) for p in norm_disp]
        week_spots = [int(x) for x in raw]
        remainder = n_spots - sum(week_spots)
        fracs = sorted([(raw[i] - week_spots[i], i) for i in range(len(raw))], reverse=True)
        for k in range(remainder):
            week_spots[fracs[k % len(fracs)][1]] += 1

        for w_idx, ws_count in enumerate(week_spots):
            if ws_count <= 0:
                continue
            week_start = schedule_start + timedelta(days=7 * w_idx)
            if week_start > camp_end:
                break
            # Slot uniqueness scoped to (_row_id, edit_duration). Each input plan
            # row gets its own slot capacity; different edits of the same row can
            # share a half-hour (multiple copy lengths in the same commercial break)
            # but the same edit can't be double-booked on the same (date, time).
            occ_key = (er.get("_row_id"), er.get("edit_duration"))
            occ = slot_occupancy.setdefault(occ_key, set())
            allocated = allocate_spots_daily(
                days_list, slot_times, slot_weights, ws_count, week_start,
                blackout_date_set, camp_end,
                week_index=w_idx,
                weekend_boost=float(prefs.weekend_boost or 1.0),
                reach_vs_frequency=float(prefs.reach_vs_frequency if prefs.reach_vs_frequency is not None else 0.5),
                used_slots=occ,
            )
            for (d, d_date, t) in allocated:
                actual_dow = DAY_ORDER[d_date.weekday()]
                if actual_dow not in days_list:
                    continue
                if d_date in blackout_date_set:
                    continue
                slot_key = (d_date, format_time(t))
                occ.add(slot_key)
                schedule_rows.append({
                    "_row_id": er["_row_id"],
                    "index_value": er["index_value"],
                    "edit_duration": er["edit_duration"],
                    "week": w_idx + 1,
                    "date": d_date.isoformat(),
                    "day": actual_dow,
                    "market": er.get("market"),
                    "genre": er.get("genre"),
                    "channel": er.get("channel"),
                    "program": er.get("program"),
                    "spot_time": format_time(t),
                    "daypart": daypart_of(t),
                })

    # Save result
    result_id = str(uuid.uuid4())
    result_doc = {
        "id": result_id,
        "plan_id": plan_id,
        "edits": [e.model_dump() for e in edits_global],
        "row_overrides": [ov.model_dump() for ov in req.row_overrides],
        "prefs": prefs.model_dump(),
        "campaign_start_monday": schedule_start.isoformat(),
        "edit_rows": edit_rows,
        "schedule_rows": schedule_rows,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.results.insert_one(result_doc)

    total_edit_fct = sum(r["final_fct"] for r in edit_rows)
    total_edit_spots = sum(r["final_spots"] for r in edit_rows)
    total_grp = sum(r["grp"] for r in edit_rows)
    total_outlay = sum(r["net_outlay"] for r in edit_rows)

    by_edit: Dict[int, Dict[str, float]] = {}
    for r in edit_rows:
        b = by_edit.setdefault(r["edit_duration"], {"duration": r["edit_duration"], "spots": 0, "fct": 0, "grp": 0, "outlay": 0})
        b["spots"] += r["final_spots"]
        b["fct"] += r["final_fct"]
        b["grp"] += r["grp"]
        b["outlay"] += r["net_outlay"]

    by_week: Dict[int, Dict[str, float]] = {}
    for r in schedule_rows:
        b = by_week.setdefault(r["week"], {"week": r["week"], "spots": 0})
        b["spots"] += 1
    # Fill missing weeks with 0
    for w in range(1, weeks + 1):
        by_week.setdefault(w, {"week": w, "spots": 0})
    by_week_list = [by_week[w] for w in sorted(by_week.keys())]

    by_channel: Dict[str, Dict[str, float]] = {}
    for r in edit_rows:
        ch = str(r.get("channel") or "Unknown")
        b = by_channel.setdefault(ch, {"channel": ch, "spots": 0, "fct": 0, "outlay": 0, "grp": 0})
        b["spots"] += r["final_spots"]
        b["fct"] += r["final_fct"]
        b["outlay"] += r["net_outlay"]
        b["grp"] += r["grp"]

    by_daypart: Dict[str, Dict[str, float]] = {}
    for r in schedule_rows:
        b = by_daypart.setdefault(r["daypart"], {"daypart": r["daypart"], "spots": 0})
        b["spots"] += 1

    return {
        "result_id": result_id,
        "edit_rows": edit_rows,
        "schedule_rows": schedule_rows,
        "campaign_start_monday": schedule_start.isoformat(),
        "campaign_start": schedule_start.isoformat(),
        "campaign_end": camp_end.isoformat(),
        "summary": {
            "total_edit_fct": round(total_edit_fct, 2),
            "total_edit_spots": total_edit_spots,
            "total_grp": round(total_grp, 3),
            "total_outlay": round(total_outlay, 2),
            "by_edit": list(by_edit.values()),
            "by_week": by_week_list,
            "by_channel": list(by_channel.values()),
            "by_daypart": list(by_daypart.values()),
        }
    }

# -------------------- Excel export in original format --------------------

def build_output_workbook(plan_doc: Dict[str, Any], result_doc: Dict[str, Any]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule Sheet"

    metadata = plan_doc.get("metadata", {})
    columns = plan_doc.get("columns", [])
    raw_rows = plan_doc.get("raw_rows", [])
    edit_rows = result_doc["edit_rows"]
    schedule_rows = result_doc["schedule_rows"]
    prefs = result_doc["prefs"]
    camp_start = datetime.fromisoformat(result_doc.get("campaign_start", result_doc["campaign_start_monday"])).date()
    # Determine weeks + hard end date; the schedule stops exactly on campaign_end (last week may be short)
    camp_end: Optional[date] = None
    if prefs.get("campaign_end"):
        try:
            camp_end = datetime.strptime(prefs["campaign_end"], "%Y-%m-%d").date()
            total_days = (camp_end - camp_start).days + 1
            weeks = max(1, math.ceil(total_days / 7))
        except Exception:
            weeks = int(prefs.get("campaign_weeks", 6))
    else:
        weeks = int(prefs.get("campaign_weeks", 6))
    if camp_end is None:
        camp_end = camp_start + timedelta(days=weeks * 7 - 1)
    # Total days is bounded by campaign_end (last week may be partial)
    n_days = (camp_end - camp_start).days + 1

    # Styling
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="002FA7")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill("solid", fgColor="E8F0FE")
    grand_fill = PatternFill("solid", fgColor="002FA7")
    grand_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    # Metadata rows
    row = 1
    for k, v in metadata.items():
        ws.cell(row=row, column=2, value=k).font = bold
        ws.cell(row=row, column=3, value=str(v))
        row += 1
    row += 1  # gap

    output_cols = list(columns)
    extra_cols = ["Edit", "Final Spots", "Final FCT", "Net Outlay Recomputed", "GRP Recomputed"]
    date_cols = [(camp_start + timedelta(days=i)) for i in range(n_days)]
    week_summary_start_label = "Weekly Spots"
    week_disp_label = "Weekly Spot Dispersion"
    week_grp_pct_label = "Weekly GRP %"

    header_row_idx = row
    for w in range(weeks):
        c_start = len(output_cols) + len(extra_cols) + 1 + w * 7
        ws.cell(row=row, column=c_start, value=f"Wk {w+1}").font = bold
    row += 1
    for i, dt in enumerate(date_cols):
        col = len(output_cols) + len(extra_cols) + 1 + i
        ws.cell(row=row, column=col, value=DAY_ORDER[dt.weekday()]).font = bold
    row += 1

    col_names_row = row
    for j, name in enumerate(output_cols):
        c = ws.cell(row=col_names_row, column=j + 1, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    for j, name in enumerate(extra_cols):
        c = ws.cell(row=col_names_row, column=len(output_cols) + j + 1, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    date_col_start = len(output_cols) + len(extra_cols) + 1
    for i, dt in enumerate(date_cols):
        c = ws.cell(row=col_names_row, column=date_col_start + i, value=dt)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.number_format = "dd-mmm"
    ws_start = date_col_start + n_days + 1
    ws.cell(row=header_row_idx, column=ws_start, value=week_summary_start_label).font = bold
    for w in range(weeks):
        c = ws.cell(row=col_names_row, column=ws_start + w, value=f"Wk {w+1}")
        c.font = header_font
        c.fill = header_fill
    disp_start = ws_start + weeks + 1
    ws.cell(row=header_row_idx, column=disp_start, value=week_disp_label).font = bold
    for w in range(weeks):
        c = ws.cell(row=col_names_row, column=disp_start + w, value=f"Wk {w+1}")
        c.font = header_font
        c.fill = header_fill
    grp_pct_start = disp_start + weeks + 1
    ws.cell(row=header_row_idx, column=grp_pct_start, value=week_grp_pct_label).font = bold
    for w in range(weeks):
        c = ws.cell(row=col_names_row, column=grp_pct_start + w, value=f"Wk {w+1}")
        c.font = header_font
        c.fill = header_fill

    row = col_names_row + 1

    raw_by_id = {r["_row_id"]: r for r in raw_rows}
    sched_index: Dict[tuple, int] = {}
    for s in schedule_rows:
        key = (s["_row_id"], s["edit_duration"], s["date"])
        sched_index[key] = sched_index.get(key, 0) + 1
    date_iso_list = [d.isoformat() for d in date_cols]

    # Attach per-spot-GRP for each row (constant per plan row)
    per_row_grp_per_spot: Dict[int, float] = {}
    for er in edit_rows:
        rid = er["_row_id"]
        if rid in per_row_grp_per_spot:
            continue
        s = safe_num(er.get("final_spots"))
        g = safe_num(er.get("grp"))
        per_row_grp_per_spot[rid] = (g / s) if s > 0 else 0

    def write_data_row(er: Dict, raw: Dict, target_row: int):
        for j, cname in enumerate(output_cols):
            v = raw.get(cname)
            ws.cell(row=target_row, column=j + 1, value=v)
        base = len(output_cols)
        ws.cell(row=target_row, column=base + 1, value=er["edit_duration"])
        ws.cell(row=target_row, column=base + 2, value=er["final_spots"])
        ws.cell(row=target_row, column=base + 3, value=er["final_fct"])
        ws.cell(row=target_row, column=base + 4, value=er["net_outlay"])
        ws.cell(row=target_row, column=base + 5, value=er["grp"])

        weekly_counts = [0] * weeks
        rid = er["_row_id"]
        for i, iso in enumerate(date_iso_list):
            n = sched_index.get((rid, er["edit_duration"], iso), 0)
            if n:
                ws.cell(row=target_row, column=date_col_start + i, value=n)
                weekly_counts[i // 7] += n
        total_week_spots = sum(weekly_counts)
        for w in range(weeks):
            ws.cell(row=target_row, column=ws_start + w, value=weekly_counts[w])
        for w in range(weeks):
            pct = (weekly_counts[w] * 100.0 / total_week_spots) if total_week_spots else 0
            ws.cell(row=target_row, column=disp_start + w, value=round(pct, 2))
        # Weekly GRP % (weekly_grp / total_row_grp * 100)
        row_grp = safe_num(er.get("grp"))
        gps = per_row_grp_per_spot.get(rid, 0)
        for w in range(weeks):
            weekly_grp = weekly_counts[w] * gps
            pct = (weekly_grp * 100.0 / row_grp) if row_grp else 0
            ws.cell(row=target_row, column=grp_pct_start + w, value=round(pct, 2))
        return weekly_counts

    def write_subtotal(label: str, group_rows: List[Dict], target_row: int, level: str = "channel"):
        # Compute totals
        spots = sum(er["final_spots"] for er in group_rows)
        fct = sum(er["final_fct"] for er in group_rows)
        outlay = sum(er["net_outlay"] for er in group_rows)
        grp = sum(er["grp"] for er in group_rows)
        # Weekly counts aggregate
        weekly_counts = [0] * weeks
        for er in group_rows:
            rid = er["_row_id"]
            for i, iso in enumerate(date_iso_list):
                n = sched_index.get((rid, er["edit_duration"], iso), 0)
                weekly_counts[i // 7] += n
        total_week_spots = sum(weekly_counts)
        row_grp_total = grp

        # Label cell + fills
        c = ws.cell(row=target_row, column=1, value=label)
        c.font = bold
        for j in range(1, len(output_cols) + len(extra_cols) + 1 + n_days + 3 * weeks + 2):
            ws.cell(row=target_row, column=j).fill = subtotal_fill if level != "grand" else grand_fill
            if level == "grand":
                ws.cell(row=target_row, column=j).font = grand_font

        base = len(output_cols)
        ws.cell(row=target_row, column=base + 2, value=spots).font = bold
        ws.cell(row=target_row, column=base + 3, value=round(fct, 2)).font = bold
        ws.cell(row=target_row, column=base + 4, value=round(outlay, 2)).font = bold
        ws.cell(row=target_row, column=base + 5, value=round(grp, 3)).font = bold
        for i, iso in enumerate(date_iso_list):
            n = 0
            for er in group_rows:
                n += sched_index.get((er["_row_id"], er["edit_duration"], iso), 0)
            if n:
                ws.cell(row=target_row, column=date_col_start + i, value=n).font = bold
        for w in range(weeks):
            ws.cell(row=target_row, column=ws_start + w, value=weekly_counts[w]).font = bold
        for w in range(weeks):
            pct = (weekly_counts[w] * 100.0 / total_week_spots) if total_week_spots else 0
            ws.cell(row=target_row, column=disp_start + w, value=round(pct, 2)).font = bold
        # Weekly GRP % for the group: sum(weekly_spots_of_each_row * gps_of_that_row)
        for w in range(weeks):
            weekly_grp = 0.0
            for er in group_rows:
                rid = er["_row_id"]
                gps = per_row_grp_per_spot.get(rid, 0)
                w_start_idx = w * 7
                w_end_idx = w_start_idx + 7
                n = 0
                for i in range(w_start_idx, min(w_end_idx, len(date_iso_list))):
                    n += sched_index.get((rid, er["edit_duration"], date_iso_list[i]), 0)
                weekly_grp += n * gps
            pct = (weekly_grp * 100.0 / row_grp_total) if row_grp_total else 0
            ws.cell(row=target_row, column=grp_pct_start + w, value=round(pct, 2)).font = bold

    # Preserve input order for market/genre/channel; edits within a row: highest duration first
    market_ord: Dict[str, int] = {}
    genre_ord: Dict[str, int] = {}
    channel_ord: Dict[str, int] = {}
    for raw in raw_rows:
        m = str(raw.get("Market") or raw.get("market") or "")
        g = str(raw.get("Genre") or raw.get("genre") or "")
        ch = str(raw.get("Channel") or raw.get("channel") or "")
        if m and m not in market_ord:
            market_ord[m] = len(market_ord)
        if g and g not in genre_ord:
            genre_ord[g] = len(genre_ord)
        if ch and ch not in channel_ord:
            channel_ord[ch] = len(channel_ord)

    def key_for(er):
        m = str(er.get("market") or "")
        g = str(er.get("genre") or "")
        ch = str(er.get("channel") or "")
        return (
            market_ord.get(m, 10_000),
            genre_ord.get(g, 10_000),
            channel_ord.get(ch, 10_000),
            er.get("_row_id", 0),
            -int(er.get("edit_duration", 0)),  # highest edit first
        )
    sorted_edits = sorted(edit_rows, key=key_for)

    # Emit with subtotals
    cur_market = None
    cur_genre = None
    cur_channel = None
    channel_rows: List[Dict] = []
    genre_rows: List[Dict] = []
    market_rows: List[Dict] = []
    grand_rows: List[Dict] = []

    def flush_channel():
        nonlocal row, channel_rows
        if channel_rows and cur_channel is not None:
            write_subtotal(f"{cur_channel} Total", channel_rows, row, level="channel")
            row += 1
        channel_rows = []

    def flush_genre():
        nonlocal row, genre_rows
        if genre_rows and cur_genre is not None:
            write_subtotal(f"{cur_market} {cur_genre} Total", genre_rows, row, level="genre")
            row += 1
        genre_rows = []

    def flush_market():
        nonlocal row, market_rows
        if market_rows and cur_market is not None:
            write_subtotal(f"{cur_market} Total", market_rows, row, level="market")
            row += 1
        market_rows = []

    for er in sorted_edits:
        m = er.get("market")
        g = er.get("genre")
        ch = er.get("channel")
        if cur_market is None:
            cur_market, cur_genre, cur_channel = m, g, ch
        else:
            if ch != cur_channel or g != cur_genre or m != cur_market:
                flush_channel()
                if g != cur_genre or m != cur_market:
                    flush_genre()
                    if m != cur_market:
                        flush_market()
                        cur_market = m
                    cur_genre = g
                cur_channel = ch

        raw = raw_by_id.get(er["_row_id"], {})
        write_data_row(er, raw, row)
        row += 1
        channel_rows.append(er)
        genre_rows.append(er)
        market_rows.append(er)
        grand_rows.append(er)

    # Final flushes
    flush_channel()
    flush_genre()
    flush_market()

    if grand_rows:
        write_subtotal("GRAND TOTAL", grand_rows, row, level="grand")
        row += 1

    # column widths
    for j in range(1, min(len(output_cols) + len(extra_cols) + 1, 50)):
        ws.column_dimensions[ws.cell(row=col_names_row, column=j).column_letter].width = 14

    # Second sheet: Edit Config + Preferences
    ws2 = wb.create_sheet("Edit Config")
    ws2.append(["Duration (s)", "Percentage"])
    for e in result_doc["edits"]:
        ws2.append([e["duration"], e["percentage"]])

    ws3 = wb.create_sheet("Preferences")
    for k, v in prefs.items():
        ws3.append([k, str(v)])

    return wb

@api_router.get("/results/{result_id}/download")
async def download_result(result_id: str):
    result_doc = await db.results.find_one({"id": result_id}, {"_id": 0})
    if not result_doc:
        raise HTTPException(status_code=404, detail="Result not found")
    plan_doc = await db.plans.find_one({"id": result_doc["plan_id"]}, {"_id": 0})
    if not plan_doc:
        raise HTTPException(status_code=404, detail="Plan not found")

    wb = build_output_workbook(plan_doc, result_doc)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="acd_schedule_{result_id[:8]}.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

# -------------------- Sessions --------------------

@api_router.post("/sessions")
async def save_session(s: SessionSave):
    session_id = str(uuid.uuid4())
    doc = {
        "id": session_id,
        "name": s.name,
        "plan_id": s.plan_id,
        "edits": [e.model_dump() for e in s.edits],
        "row_overrides": [ov.model_dump() for ov in s.row_overrides],
        "prefs": s.prefs.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.sessions.insert_one(doc)
    return {"id": session_id, "name": s.name}

@api_router.get("/sessions")
async def list_sessions():
    items = await db.sessions.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return items

@api_router.get("/sessions/{sid}")
async def get_session(sid: str):
    doc = await db.sessions.find_one({"id": sid}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Session not found")
    return doc

@api_router.delete("/sessions/{sid}")
async def delete_session(sid: str):
    await db.sessions.delete_one({"id": sid})
    return {"ok": True}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
